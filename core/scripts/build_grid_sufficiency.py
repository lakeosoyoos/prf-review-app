"""Config-driven Grid Acreage Sufficiency deliverable (PRF review Stage 6).

Generalizes the per-account builder: reads account_config.yaml, classifies parcel owners with the
roster (name_match.OwnerClassifier), resolves supporting documents with the auto-built lessor index
(doc_index.DocIndex), parses legals/permits with trs_match, and writes a self-contained folder:

    <output_folder>/
        <output_title>.xlsx          (Summary · Grid Sufficiency · All Locations)
        Supporting Documents/         (every linked cert/permit/bundle + fee parcel-record sheets)
        _README.txt

STANDARD: for each PRF grid, insured-controlled acreage (leases + ownership + permits in-grid) >=
reported insured acres. Each location is matched to an in-grid source; precise legal-description
match is the goal, not required. ALL links are relative + URL-encoded (folder is portable).

    python build_grid_sufficiency.py account_config.yaml
"""
import os, re, sys, json, glob, pickle, shutil, collections, urllib.parse, argparse
import yaml, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader, PdfWriter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from name_match import OwnerClassifier, _norm
from doc_index import DocIndex
import trs_match


# --------------------------------------------------------------------------- helpers
def num(x):
    try:
        return float(str(x).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def seckey(legal):
    """Parse (T,R,S) from an FSA 'Legal' cell like '... T31N R25E Sec 33'."""
    s = str(legal or "")
    mT, mR, mS = re.search(r"(\d+)\s*N", s), re.search(r"(\d+)\s*E", s), re.search(r"(\d+)\D*$", s)
    return (int(mT[1]), int(mR[1]), int(mS[1])) if (mT and mR and mS) else None


def resolve_pkt(s, packet_glob):
    if "{PKT}" in s:
        hits = sorted(glob.glob(packet_glob)) if packet_glob else []
        if not hits:
            return None
        return s.replace("{PKT}", hits[-1])
    return s


def main(config_path, flat=False, out_override=None):
    cfg = yaml.safe_load(open(config_path))
    D = cfg["deliverable"]
    pol = (cfg.get("policies") or [{}])[0]
    county_label = cfg.get("county_label", "")
    FOLDER = out_override or D["output_folder"] + ("_FLAT" if flat else "")
    # flat = PDFs sit beside the workbook (one level) so relative links survive a flat cloud upload;
    # otherwise PDFs live in a 'Supporting Documents' subfolder (tidier for local/zipped use).
    DOCS = FOLDER if flat else os.path.join(FOLDER, "Supporting_Documents")
    link_prefix = "" if flat else "Supporting_Documents/"
    # Local Excel hyperlinks must NOT be percent-encoded and must avoid characters Windows/Excel
    # trip on (spaces, parens, &, em-dash). Sanitize every on-disk doc name to plain ASCII and link
    # to that exact name — no urllib.quote. linkname() is memoized + collision-safe.
    _safe_map, _safe_used = {}, set()

    def linkname(fn):
        if fn in _safe_map:
            return _safe_map[fn]
        base, ext = os.path.splitext(fn)
        b = base.replace("—", "-").replace("–", "-").replace("&", "and")
        b = re.sub(r"[^A-Za-z0-9._-]+", "_", b)
        b = re.sub(r"_+", "_", b).strip("_") or "doc"
        cand, i = b + ext.lower(), 2
        while cand in _safe_used:
            cand, i = f"{b}_{i}{ext.lower()}", i + 1
        _safe_used.add(cand); _safe_map[fn] = cand
        return cand
    packet_glob = D.get("packet_glob")
    src_dirs = [d for d in (resolve_pkt(x, packet_glob) for x in D.get("doc_source_dirs", [])) if d and os.path.isdir(d)]
    extra_docs = {k: v for k, v in (D.get("extra_docs") or {}).items()}

    oc = OwnerClassifier.from_config(config_path)

    # ---- parcel layer: section -> controlled owners / insured fee PINs ----------------
    parcel_pkl = (cfg.get("data_sources", {}).get("parcel_layers") or {})
    pkl_path = next(iter(parcel_pkl.values())) if parcel_pkl else "extracted/parcels.pkl"
    meta = pickle.load(open(pkl_path, "rb"))
    trs_owners = collections.defaultdict(collections.Counter)   # section -> Counter(raw owner string)
    trs_feepins = collections.defaultdict(list)                 # section -> [insured fee PINs]
    sec_parcels = collections.defaultdict(list)                 # section -> [insured fee parcel attrs]
    for a in meta["attrs"]:
        k = trs_match.parse_map(a.get("map"))
        owner = a.get("owner") or ""
        if not k:
            continue
        kind, _ = oc.classify(owner)
        if kind == "fee_insured" and str(a.get("PIN", "")).isdigit():
            trs_feepins[k].append(str(a["PIN"]))
            sec_parcels[k].append(a)
        elif kind in ("lessor", "fee_related", "gov_lease", "tribal", "private_tag"):
            trs_owners[k][owner] += 1

    # ---- recorded permits/tribal leases: section -> instrument file -------------------
    pub_trs = {}
    AGENCY_KW = [("WDFW", "WA WDFW (state lease)"), ("FISH", "WA WDFW (state lease)"),
                 ("FOREST", "US Forest Service (permit)"), ("USFS", "US Forest Service (permit)"),
                 ("BLM", "US BLM (federal permit)"), ("INTERIOR", "US BLM (federal permit)"),
                 ("DNR", "WA DNR (state lease)"), ("NATURAL RESOURCES", "WA DNR (state lease)")]
    typed = collections.defaultdict(list)   # agency label -> [recorded filenames]
    for src in (D.get("recorded_sources") or []):
        if not os.path.exists(src):
            continue
        data = json.load(open(src))
        items = data if isinstance(data, list) else list(data.values())
        for it in items:
            f = it.get("file", "")
            legs = it.get("legal_descriptions") or []
            for leg in legs:
                for trs in trs_match.parse_prose(leg):
                    if trs and trs[2] is not None and f:
                        pub_trs[trs] = f
            blob = (str(it.get("term", "")) + " " + str(it.get("lessor", ""))).upper()
            lab = next((l for kw, l in AGENCY_KW if kw in blob), "WA DNR (state lease)")
            if f:
                typed[lab].append(f)

    # ---- find sources + build agency bundles ------------------------------------------
    def find_src(fn):
        if fn in extra_docs and os.path.exists(extra_docs[fn]):
            return extra_docs[fn]
        for d in src_dirs:
            for root, _, files in os.walk(d):
                if fn in files:
                    return os.path.join(root, fn)
        return None

    if os.path.exists(FOLDER):
        shutil.rmtree(FOLDER)
    os.makedirs(DOCS)

    def merge_bundle(out, files):
        w = PdfWriter(); n = 0
        for fn in files:
            s = find_src(fn)
            if not s:
                continue
            try:
                for p in PdfReader(s).pages:
                    w.add_page(p)
                n += 1
            except Exception:
                pass
        if n:
            with open(os.path.join(DOCS, linkname(out)), "wb") as fh:
                w.write(fh)
        return out if n else None

    bundle_out = {}            # agency label -> bundle filename (if built)
    for lab, spec in (D.get("agency_bundles") or {}).items():
        out = spec.get("out")
        files = typed.get(lab, []) + (spec.get("extra") or [])
        built = merge_bundle(out, files)
        if built:
            bundle_out[lab] = built

    # ---- auto doc index ----------------------------------------------------------------
    di = DocIndex.build(src_dirs, config=config_path, oc=oc,
                        agency_bundles={lab: fn for lab, fn in bundle_out.items()})
    di.write(os.path.join(os.path.dirname(pkl_path) or ".", "doc_index.json"))

    # ---- locations from the field-verification workbook --------------------------------
    fv = openpyxl.load_workbook(D["fv_workbook"], data_only=True)[D.get("fv_sheet", "Field Verification")]
    H = [str(c.value) for c in fv[1]]; I = {h: i for i, h in enumerate(H)}
    col = lambda *names: next((I[n] for n in names if n in I), None)
    c_fsn, c_fld, c_tr = col("FSN"), col("FSA Field #", "Field"), col("Tract")
    c_ac = col("Field Reported Acres", "Reported Acres")
    c_leg, c_grid = col("Legal"), col("Assigned To", "Grid")
    c_ctl = col("Control source (section parcels)", "Control source")
    c_les = col("Lessor (party leasing to Gebbers Cattle LP)", "Lessor")
    c_flag = col("Flag")

    overrides = D.get("manual_overrides") or []

    def override_for(fsn, fld):
        for o in overrides:
            if str(o.get("fsn")) == fsn and (not o.get("fields") or fld in [str(x) for x in o["fields"]]):
                return o
        return None

    def status_of(fsn, fld, flag):
        o = override_for(fsn, fld)
        if o and o.get("status"):
            return o["status"]
        f = (flag or "").lower()
        if "lessor not identified" in f or "legal not locatable" in f:
            return "DOCUMENT NEEDED"
        return "SUPPORTED"

    rows = [r for r in fv.iter_rows(min_row=2, values_only=True) if r[c_fsn] not in (None, "")]
    loc, needed = [], set()
    for r in rows:
        fsn, fld = str(r[c_fsn]), str(r[c_fld])
        k = seckey(r[c_leg])
        les_raw = str(r[c_les] or "") if c_les is not None else ""
        ctl_raw = str(r[c_ctl] or "") if c_ctl is not None else ""
        docs = []
        for owner, _ in (trs_owners.get(k) or collections.Counter()).most_common():
            for d in di.lookup(owner, trs=k, recorded_by_trs=pub_trs):
                if d and d not in docs:
                    docs.append(d)
        # FV-row fallback: when the parcel layer has no owner for this section (out-of-county /
        # public land), use the row's OWN Lessor determination to resolve a document.
        if not docs and les_raw:
            for d in di.lookup(les_raw, trs=k, recorded_by_trs=pub_trs):
                if d and d not in docs:
                    docs.append(d)
        ov = override_for(fsn, fld)
        if ov and ov.get("insert_doc") and ov["insert_doc"] not in docs:
            docs.insert(0, ov["insert_doc"])
            extra_docs.setdefault(ov["insert_doc"], extra_docs.get(ov["insert_doc"]))
        pins = trs_feepins.get(k) or []
        feenote = ("Insured fee: " + ", ".join(pins[:8]) + (f"  (+{len(pins) - 8} more)" if len(pins) > 8 else "")) if pins else ""
        for d in docs:
            needed.add(d)
        loc.append({"grid": str(r[c_grid]).split()[0], "fsn": fsn, "tract": r[c_tr], "field": fld,
                    "ac": num(r[c_ac]), "legal": r[c_leg],
                    "ctl": str(r[c_ctl] or "")[:46] if c_ctl is not None else "",
                    "lessor": str(r[c_les] or "")[:30] if c_les is not None else "",
                    "status": status_of(fsn, fld, str(r[c_flag] or "") if c_flag is not None else ""),
                    "note": (ov.get("note", "") if ov else ""), "docs": docs, "fee": feenote, "k": k})

    # fallbacks: (1) agency bundle by control/lessor keyword, (2) grid-dominant doc
    for d in loc:
        if d["docs"]:
            continue
        blob = (d["ctl"] + " " + d["lessor"]).upper()
        lab = next((l for kw, l in AGENCY_KW if kw in blob), None)
        cand = bundle_out.get(lab) if lab else None
        if cand:
            d["docs"] = [cand]; needed.add(cand)
    gd = collections.defaultdict(collections.Counter)
    for d in loc:
        for x in d["docs"]:
            gd[d["grid"]][x] += 1
    for d in loc:
        if not d["docs"] and gd[d["grid"]]:
            d["docs"] = [gd[d["grid"]].most_common(1)[0][0]]; needed.add(d["docs"][0])

    # ---- local assessor-style parcel-record sheets for insured fee ground (no network) -
    from reportlab.lib.pagesizes import letter as RLET
    from reportlab.lib.units import inch as RIN
    from reportlab.lib import colors as RC
    from reportlab.platypus import SimpleDocTemplate, Paragraph as RP, Spacer as RS, Table as RT, TableStyle as RTS
    from reportlab.lib.styles import getSampleStyleSheet
    _ss = getSampleStyleSheet(); _cell = _ss["Normal"]; _cell.fontSize = 8; _cell.leading = 10
    insured_name = pol.get("insured", cfg.get("account_name", "Insured"))

    def section_record(k):
        T, R, S = k
        fn = f"Owned fee parcel records — T{T}N R{R}E Sec {S}.pdf"
        path = os.path.join(DOCS, linkname(fn))
        if os.path.exists(path):
            return fn
        ps = sorted(sec_parcels.get(k, []), key=lambda a: str(a.get("PIN", "")))
        doc = SimpleDocTemplate(path, pagesize=RLET, topMargin=0.6 * RIN, bottomMargin=0.5 * RIN,
                                leftMargin=0.6 * RIN, rightMargin=0.6 * RIN)
        E = [RP(f"<b>{insured_name} — Owned (Fee) Parcel Records</b>", _ss["Title"]),
             RP(f"Section {S}, Township {T} North, Range {R} East — {county_label} &nbsp;·&nbsp; source: county "
                f"assessor parcel layer (local). Ownership record (not the recorded deed; the deed is in the "
                f"County Auditor's recording system).", _ss["Normal"]), RS(1, 8)]
        tbl = [[RP(f"<b>{h}</b>", _cell) for h in ["Parcel # (PIN)", "Owner", "Situs", "Deeded ac", "Legal description"]]]
        for a in ps:
            tbl.append([RP(str(a.get("PIN", "")), _cell), RP(str(a.get("owner", "")), _cell),
                        RP(str(a.get("situs", "")), _cell), RP(f'{num(a.get("deeded_acres", 0)):.2f}', _cell),
                        RP(str(a.get("legal", ""))[:120], _cell)])
        t = RT(tbl, colWidths=[1.1 * RIN, 1.7 * RIN, 1.5 * RIN, 0.7 * RIN, 2.4 * RIN], repeatRows=1)
        t.setStyle(RTS([("BACKGROUND", (0, 0), (-1, 0), RC.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), RC.white),
                        ("GRID", (0, 0), (-1, -1), 0.4, RC.HexColor("#BFBFBF")), ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [RC.white, RC.HexColor("#F2F6FA")])]))
        E.append(t); doc.build(E)
        return fn

    for d in loc:
        if d["fee"] and d["k"] in sec_parcels:
            fn = section_record(d["k"])
            if fn not in d["docs"]:
                d["docs"].append(fn)

    # ---- copy referenced docs into the folder -----------------------------------------
    missing = []
    for fn in sorted(needed):
        if os.path.exists(os.path.join(DOCS, linkname(fn))):
            continue
        s = find_src(fn)
        if s:
            shutil.copy2(s, os.path.join(DOCS, linkname(fn)))
        else:
            missing.append(fn)

    maxdocs = max((len(d["docs"]) for d in loc), default=1)
    nolink = sum(1 for d in loc if not d["docs"] and not d["fee"])
    broken = sum(1 for d in loc for fn in d["docs"] if not os.path.exists(os.path.join(DOCS, linkname(fn))))

    # ---- workbook ---------------------------------------------------------------------
    NAVY = PatternFill("solid", fgColor="1F4E78"); GRN = PatternFill("solid", fgColor="D9EAD3")
    YEL = PatternFill("solid", fgColor="FFF2CC"); BLU = PatternFill("solid", fgColor="DDEBF7"); HEAD = PatternFill("solid", fgColor="EEF2F7")
    WH = Font(bold=True, color="FFFFFF"); B = Font(bold=True); TITLE = Font(bold=True, size=14, color="1F4E78")
    LINK = Font(color="0563C1", underline="single")
    WRAP = Alignment(wrap_text=True, vertical="top"); TOP = Alignment(vertical="top")
    thin = Side(style="thin", color="BFBFBF"); BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    G = collections.defaultdict(lambda: {"rep": 0.0, "n": 0, "dn": 0.0, "ndn": 0, "flag": 0})
    for d in loc:
        a = G[d["grid"]]; a["rep"] += d["ac"]; a["n"] += 1
        if d["status"] == "DOCUMENT NEEDED":
            a["dn"] += d["ac"]; a["ndn"] += 1
        if "flagged" in d["status"]:
            a["flag"] += 1
    tot = sum(a["rep"] for a in G.values()); tot_dn = sum(a["dn"] for a in G.values())
    cov_pct = 100 * (tot - tot_dn) / tot if tot else 0

    s = wb.create_sheet("Summary"); s["A1"] = D.get("output_title", "Grid Acreage Sufficiency"); s["A1"].font = TITLE
    lines = [
        f"Policy {pol.get('policy_number', '')} · {county_label} · {len(loc)} field locations. Self-contained — PDFs in 'Supporting Documents'.",
        "'All Locations' links EACH location to EVERY distinct supporting document for its section (one column per doc).",
        "Insured-owned (fee) ground has no lease PDF — its county parcel #s are in the 'Owned (fee)' column, plus a generated parcel-record sheet.", "",
        ("STANDARD: per PRF grid, insured-controlled acreage (leases+ownership+permits in-grid) >= reported insured acres.", True),
        (f"RESULT: {tot - tot_dn:,.0f} / {tot:,.0f} reported acres ({cov_pct:.1f}%) covered in-grid; "
         f"{sum(1 for g in G if G[g]['dn'] < 0.5)} of {len(G)} grids fully covered.", True), "",
        ("Precise legal-description match is the goal but not required; same-grid acreage sufficiency is the standard.", False),
    ]
    for i, ln in enumerate(lines, 1):
        t, bd = (ln if isinstance(ln, tuple) else (ln, False)); c = s.cell(i + 1, 1, t)
        if bd:
            c.font = B
    s.column_dimensions["A"].width = 124

    g1 = wb.create_sheet("Grid Sufficiency")
    g1.append(["Grid", "Reported ac", "Covered in-grid ac", "Doc-needed ac", "# loc", "# flagged", "Status"])
    for c in g1[1]:
        c.fill = NAVY; c.font = WH; c.alignment = WRAP; c.border = BORD
    for g in sorted(G, key=lambda k: -G[k]["rep"]):
        a = G[g]; dn = round(a["dn"], 1)
        st = ("SUPPORTED — all in-grid" if dn < 0.5 and a["flag"] == 0 else
              "SUPPORTED — 1+ source flagged for review" if dn < 0.5 else
              f"{a['ndn']} location(s) need a document ({dn} ac)")
        g1.append([g, round(a["rep"], 1), round(a["rep"] - a["dn"], 1), dn, a["n"], a["flag"], st])
        fill = GRN if dn < 0.5 and a["flag"] == 0 else (BLU if dn < 0.5 else YEL)
        for c in g1[g1.max_row]:
            c.fill = fill; c.border = BORD; c.alignment = TOP
    g1.append(["TOTAL", round(tot, 1), round(tot - tot_dn, 1), round(tot_dn, 1), len(loc),
               sum(a["flag"] for a in G.values()), f"{cov_pct:.1f}% covered in-grid ({len(G)} grids)"])
    for c in g1[g1.max_row]:
        c.font = B; c.fill = HEAD; c.border = BORD
    g1.freeze_panes = "A2"
    for i, w in enumerate([10, 13, 18, 13, 7, 12, 56], 1):
        g1.column_dimensions[get_column_letter(i)].width = w

    g3 = wb.create_sheet("All Locations")
    base = ["Grid", "FSN", "Tract", "Field", "Reported ac", "Legal", "Control source (section)",
            "Lessor / owner", "Status", "Owned (fee) — county parcel #s", "Note"]
    doccols = [f"Supporting doc {i + 1} (click)" for i in range(maxdocs)]
    g3.append(base + doccols)
    for c in g3[1]:
        c.fill = NAVY; c.font = WH; c.alignment = WRAP
    for d in sorted(loc, key=lambda x: (x["grid"], -x["ac"])):
        g3.append([d["grid"], d["fsn"], d["tract"], d["field"], round(d["ac"], 2), d["legal"], d["ctl"],
                   d["lessor"], d["status"], d["fee"], d["note"]] + [None] * maxdocs)
        rr = g3.max_row
        if d["status"] != "SUPPORTED":
            for c in g3[rr][:len(base)]:
                c.fill = YEL if d["status"] == "DOCUMENT NEEDED" else BLU
        for j, fn in enumerate(d["docs"]):
            cell = g3.cell(rr, len(base) + 1 + j); cell.value = linkname(fn)
            cell.hyperlink = link_prefix + linkname(fn); cell.font = LINK
    g3.freeze_panes = "A2"
    for i, w in enumerate([8, 7, 7, 6, 10, 14, 38, 26, 24, 30, 40] + [38] * maxdocs, 1):
        g3.column_dimensions[get_column_letter(i)].width = w

    safe_title = re.sub(r"_+", "_", re.sub(r"[^A-Za-z0-9._-]+", "_",
                        D.get("output_title", "Grid Acreage Sufficiency").replace("—", "-"))).strip("_")
    out_xlsx = os.path.join(FOLDER, safe_title + ".xlsx")
    wb.save(out_xlsx)
    layout = ("LAYOUT: this is the FLAT version — the workbook and every PDF are in ONE folder (no "
              "subfolder). To use the clickable links, download ALL files into a single folder and open "
              "the workbook in desktop Excel. (Clicking links inside a browser/cloud preview will not open "
              "the PDFs — that is a cloud-viewer limitation.)\n\n"
              if flat else
              "LAYOUT: PDFs are in the 'Supporting Documents' subfolder; keep the folder intact so the "
              "relative links resolve.\n\n")
    with open(os.path.join(FOLDER, "_README.txt"), "w") as f:
        f.write(f"{D.get('output_title')}\n\n{layout}"
                "'All Locations' links each location to every distinct supporting document for its section "
                "(one column per document). Insured-owned (fee) ground has no lease PDF — its county parcel "
                "numbers are in the 'Owned (fee)' column and a generated parcel-record sheet is linked. "
                "Public-land ground not tied to a single permit links to an agency bundle.\n")

    print(f"WROTE {out_xlsx}")
    print(f"  locations: {len(loc)} | grids: {len(G)} | reported {tot:,.0f} ac | covered {tot - tot_dn:,.0f} ac ({cov_pct:.1f}%)")
    print(f"  doc columns: {maxdocs} | supporting files: {len(os.listdir(DOCS))}")
    print(f"  rows no-link&no-fee: {nolink} | broken links: {broken} | missing source files: {len(missing)}")
    if missing:
        print("  MISSING (referenced, not found):", missing[:10])


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Build the Grid Acreage Sufficiency deliverable from account_config.yaml")
    ap.add_argument("config", help="path to account_config.yaml")
    ap.add_argument("--flat", action="store_true",
                    help="put PDFs beside the workbook (one folder, no subfolder) for flat cloud upload")
    ap.add_argument("--out", help="override output folder path")
    a = ap.parse_args()
    main(a.config, flat=a.flat, out_override=a.out)
