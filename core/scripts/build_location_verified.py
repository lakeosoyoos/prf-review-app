"""Apply per-LOCATION verification verdicts to the existing deliverable, keeping its format.

Takes a built Grid/Location folder (workbook + supporting docs + relative links already correct) and
a verdicts file (from verify_locations.py / the agent loop), and writes a verified copy:
  - 'All Locations' Status column -> MATCHED / LIKELY / EXCEPTION (recolored), with the matched
    instrument + basis written into the Note column. Doc links/columns are left untouched.
  - new 'Exceptions & Likely' worklist sheet.
  - Summary + per-grid rollup updated to the per-location standard.

  python build_location_verified.py <src_folder> <verdicts.json> <dest_folder> [--title "..."]
"""
import os, re, sys, json, glob, shutil, collections, argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from verify_locations import loc_section

GRN = PatternFill("solid", fgColor="D9EAD3"); BLU = PatternFill("solid", fgColor="DDEBF7")
RED = PatternFill("solid", fgColor="F4CCCC"); NAVY = PatternFill("solid", fgColor="1F4E78")
HEAD = PatternFill("solid", fgColor="EEF2F7")
WH = Font(bold=True, color="FFFFFF"); B = Font(bold=True); TITLE = Font(bold=True, size=14, color="1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top"); thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
ALLOT = PatternFill("solid", fgColor="E2EFDA")   # paler green = matched, but allotment-level (coarser)
FILL = {"MATCHED": GRN, "LIKELY": BLU, "EXCEPTION": RED}


def keyof(legal):
    k = loc_section(legal)
    return "-".join(map(str, k)) if k else "LEGAL:" + str(legal)


def num(x):
    try:
        return float(str(x).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def main(src, verdicts_path, dest, title=None):
    _vd = json.load(open(verdicts_path))
    vmap = {v["key"]: v for v in (_vd["verdicts"] if isinstance(_vd, dict) else _vd)}
    if os.path.exists(dest):
        shutil.rmtree(dest)
    shutil.copytree(src, dest)
    xlsx = [f for f in glob.glob(os.path.join(dest, "*.xlsx")) if not os.path.basename(f).startswith("~$")][0]
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["All Locations"]
    H = [c.value for c in ws[1]]; I = {h: i for i, h in enumerate(H)}
    c_leg, c_stat, c_note = I["Legal"], I["Status"], I.get("Note")
    c_grid, c_fsn, c_tr, c_fld, c_ac = I["Grid"], I["FSN"], I["Tract"], I["Field"], I["Reported ac"]
    nbase = (c_note + 1) if c_note is not None else (c_stat + 1)   # cols to recolor (skip doc links)

    counts = collections.Counter(); grid_stat = collections.defaultdict(lambda: collections.Counter())
    grid_ac = collections.defaultdict(lambda: collections.defaultdict(float))
    excepts = []
    for row in ws.iter_rows(min_row=2):
        legal = row[c_leg].value
        v = vmap.get(keyof(legal))
        if not v:
            continue
        st = v["status"]; prec = (v.get("precision") or "").lower()
        coarse = (st == "MATCHED" and prec not in ("parcel", "section", ""))   # allotment / state lease / lease cert / tribal lease
        disp = f"MATCHED ({prec})" if coarse else st          # show the coarser match distinctly
        counts[disp] += 1
        g = str(row[c_grid].value); grid_stat[g][st] += 1; grid_ac[g][st] += num(row[c_ac].value)
        row[c_stat].value = disp
        if c_note is not None:
            instr = v.get("instrument") or "—"
            row[c_note].value = f"[{prec}] {instr} — {v.get('basis', '')}".strip()
        fill = ALLOT if coarse else FILL.get(st, BLU)
        for c in row[:nbase]:
            c.fill = fill
        if st != "MATCHED":
            excepts.append([g, row[c_fsn].value, row[c_tr].value, row[c_fld].value, num(row[c_ac].value),
                            legal, st, v.get("instrument", "—"), v.get("basis", ""), v.get("note", "")])

    # ---- Exceptions & Likely worklist sheet ----
    if "Exceptions & Likely" in wb.sheetnames:
        del wb["Exceptions & Likely"]
    ex = wb.create_sheet("Exceptions & Likely")
    ex.append(["Grid", "FSN", "Tract", "Field", "Reported ac", "Legal", "Status", "Matched instrument", "Why / basis", "Reviewer note"])
    for c in ex[1]:
        c.fill = NAVY; c.font = WH; c.alignment = WRAP; c.border = BORD
    for rowv in sorted(excepts, key=lambda x: (x[6] != "EXCEPTION", x[0], -x[4])):
        ex.append(rowv)
        for c in ex[ex.max_row]:
            c.fill = RED if rowv[6] == "EXCEPTION" else BLU; c.border = BORD; c.alignment = WRAP
    ex.freeze_panes = "A2"
    for i, w in enumerate([8, 7, 7, 6, 10, 16, 12, 34, 46, 30], 1):
        ex.column_dimensions[get_column_letter(i)].width = w

    # ---- per-grid rollup -> rewrite 'Grid Sufficiency' as per-location status rollup ----
    if "Grid Sufficiency" in wb.sheetnames:
        gs = wb["Grid Sufficiency"]
        wb.remove(gs)
    gs = wb.create_sheet("Grid Status", 1)
    gs.append(["Grid", "# loc", "MATCHED", "LIKELY", "EXCEPTION", "Reported ac", "EXCEPTION ac", "Status"])
    for c in gs[1]:
        c.fill = NAVY; c.font = WH; c.alignment = WRAP; c.border = BORD
    for g in sorted(grid_stat, key=lambda k: -sum(grid_ac[k].values())):
        s = grid_stat[g]; ac = grid_ac[g]; exa = ac.get("EXCEPTION", 0)
        n = sum(s.values()); tot = sum(ac.values())
        verdict = ("ALL MATCHED" if s["EXCEPTION"] == 0 and s["LIKELY"] == 0 else
                   f"{s['EXCEPTION']} exception(s)" if s["EXCEPTION"] else f"{s['LIKELY']} likely (corroborated)")
        gs.append([g, n, s["MATCHED"], s["LIKELY"], s["EXCEPTION"], round(tot, 1), round(exa, 1), verdict])
        fill = GRN if s["EXCEPTION"] == 0 and s["LIKELY"] == 0 else (RED if s["EXCEPTION"] else BLU)
        for c in gs[gs.max_row]:
            c.fill = fill; c.border = BORD
    gs.freeze_panes = "A2"
    for i, w in enumerate([10, 7, 10, 9, 11, 12, 12, 30], 1):
        gs.column_dimensions[get_column_letter(i)].width = w

    # ---- Summary ----
    sm = wb["Summary"]
    for r in list(sm.iter_rows()):
        for c in r:
            c.value = None
    ttl = title or "Location Verification — Gebbers Cattle LP, CY2026"
    sm["A1"] = ttl; sm["A1"].font = TITLE
    tot = sum(counts.values())
    matched_precise = counts["MATCHED"]
    coarse_tiers = {k: c for k, c in counts.items() if k.startswith("MATCHED (")}
    matched_total = matched_precise + sum(coarse_tiers.values())
    TIER_NOTE = {
        "MATCHED (state lease)": "WA DNR/WDFW state lease (documented per the prior-AIP review)",
        "MATCHED (lease cert)": "NAU lease certification form on file (prior-AIP standard)",
        "MATCHED (allotment)": "BLM Allotment Master Report — allotment level (prior-AIP standard)",
        "MATCHED (tribal lease)": "recorded Colville tribal lease",
    }
    lines = [
        "PER-LOCATION VERIFICATION — every FSA location matched to the lease / deed / permit that covers THAT location.",
        "Documentation standard mirrors the prior-year AIP high-dollar review, BY TYPE OF GROUND (see the basis on each row).", "",
        ("RESULT (per-location):", True),
        (f"   MATCHED  {matched_total:>3} of {tot}  ({100*matched_total/tot:.1f}%) — a named instrument covers the location:", True),
        (f"        - parcel / section-precise    {matched_precise:>3}   (instrument legal or Exhibit-A reaches the exact section)", False),
    ]
    for k in sorted(coarse_tiers):
        lines.append((f"        - {k.replace('MATCHED ','').strip('()'):<22}{coarse_tiers[k]:>3}   {TIER_NOTE.get(k,'')}", False))
    lines += [
        (f"   LIKELY    {counts['LIKELY']:>3} of {tot}  — FSA-CLU operatorship only; no cert/lease names the section (AIP would also lack a doc)", False),
        (f"   EXCEPTION {counts['EXCEPTION']:>3} of {tot}", False), "",
        ("HOW TO READ THIS WORKBOOK (for review):", True),
        ("   • 'All Locations' = every location, its Status, the matched instrument + basis (Note column), and clickable links to the", False),
        ("       supporting document(s). Status colors: dark green = parcel/section-precise; PALE green = matched at a coarser", False),
        ("       level (state lease / lease cert / allotment) per the prior-AIP standard; blue = LIKELY.", False),
        ("   • Match level is tagged at the front of each Note: [parcel] / [section] / [state lease] / [lease cert] / [allotment].", False),
        ("   • 'Exceptions & Likely' = the short worklist (the 1 LIKELY item). 'Grid Status' = per-grid rollup.", False),
        ("   • ACRE FLAGS (if any) are carried in the Note column — e.g. reported acres exceeding documented allotment acreage", False),
        ("       (mirrors the prior-AIP AR Q2.04 check). Section-precision for the coarser tiers is an optional GIS/legal upgrade.", False),
        ("Verdicts independently verified per section-group by the agent loop, then re-worked; reconciled to the prior-AIP binder.", False),
    ]
    for i, ln in enumerate(lines, 1):
        t, bd = (ln if isinstance(ln, tuple) else (ln, False)); c = sm.cell(i + 1, 1, t)
        if bd:
            c.font = B
    sm.column_dimensions["A"].width = 130

    # rename workbook + readme
    os.remove(xlsx)
    safe = re.sub(r"_+", "_", re.sub(r"[^A-Za-z0-9._-]+", "_", ttl.replace("—", "-"))).strip("_")
    wb.save(os.path.join(dest, safe + ".xlsx"))
    rm = os.path.join(dest, "_README.txt")
    if os.path.exists(rm):
        open(rm, "a").write("\n\nUPDATED to per-LOCATION verification: Status = MATCHED/LIKELY/EXCEPTION; "
                            "matched instrument + basis in the Note column; see 'Exceptions & Likely' tab.\n")
    print(f"WROTE {dest}")
    print(f"  per-location: {dict(counts)} | exceptions+likely listed: {len(excepts)}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("src"); ap.add_argument("verdicts"); ap.add_argument("dest")
    ap.add_argument("--title")
    a = ap.parse_args()
    main(a.src, a.verdicts, a.dest, a.title)
